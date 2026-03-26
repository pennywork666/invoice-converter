from __future__ import annotations

import base64
import re
from datetime import datetime
from io import BytesIO
from pathlib import Path

import streamlit as st
import streamlit.components.v1 as components


OUTPUT_COLUMNS = [
    "Employee Name",
    "Week Ending",
    "Invoice Amount",
    "System URN",
    "Invoice Number",
    "Invoice Date",
    "Due Date",
    "Project ID",
]


def clean_text(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def parse_date(value):
    if not value:
        return None

    value = clean_text(value)
    date_formats = [
        "%m/%d/%Y",
        "%m-%d-%Y",
        "%Y-%m-%d",
        "%m/%d/%y",
        "%m-%d-%y",
    ]

    for fmt in date_formats:
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    return None


def parse_currency(value):
    if not value:
        return None

    cleaned = str(value).replace("$", "").replace(",", "").strip()
    match = re.search(r"-?\d+(?:\.\d+)?", cleaned)
    if not match:
        return None
    return float(match.group())


def parse_system_urn_from_filename(filename):
    stem = Path(filename).stem
    match = re.search(r"(\d{8}_\d{6})", stem)
    if match:
        return f"PDF_IFR_{match.group(1)}"
    return f"PDF_IFR_{stem}"


def extract_text_from_pdf(pdf_bytes):
    import pdfplumber

    all_text = []
    with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            all_text.append(text)
    return "\n".join(all_text)


def extract_by_label(text, label_patterns, stop_labels=None):
    stop_labels = stop_labels or [
        "Invoice Date",
        "Due Date",
        "Invoice Number",
        "Project ID",
        "Amount Due",
        "Bill To",
        "Employee Name",
        "Week Ending",
    ]

    for label in label_patterns:
        pattern = rf"{label}\s*[:\-]?\s*([^\n\r]+)"
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            value = clean_text(match.group(1))
            for stop in stop_labels:
                value = re.split(
                    rf"\b{re.escape(stop)}\b",
                    value,
                    flags=re.IGNORECASE,
                )[0].strip()
            return value

    return ""


def group_words_into_rows(words, y_tolerance=3):
    if not words:
        return []

    words_sorted = sorted(words, key=lambda word: (round(word["top"], 1), word["x0"]))
    rows = []

    for word in words_sorted:
        placed = False
        for row in rows:
            if abs(word["top"] - row["top"]) <= y_tolerance:
                row["words"].append(word)
                row["tops"].append(word["top"])
                placed = True
                break
        if not placed:
            rows.append(
                {
                    "top": word["top"],
                    "tops": [word["top"]],
                    "words": [word],
                }
            )

    for row in rows:
        row["words"] = sorted(row["words"], key=lambda item: item["x0"])
        row["top"] = sum(row["tops"]) / len(row["tops"])

    return sorted(rows, key=lambda row: row["top"])


def find_header_positions(words):
    employee_header = None
    week_header = None

    for index in range(len(words)):
        first = clean_text(words[index].get("text", "")).lower()

        if first == "employee" and index + 1 < len(words):
            second = clean_text(words[index + 1].get("text", "")).lower()
            if second == "name" and abs(words[index]["top"] - words[index + 1]["top"]) <= 3:
                employee_header = {
                    "x0": min(words[index]["x0"], words[index + 1]["x0"]),
                    "x1": max(words[index]["x1"], words[index + 1]["x1"]),
                    "top": min(words[index]["top"], words[index + 1]["top"]),
                    "bottom": max(words[index]["bottom"], words[index + 1]["bottom"]),
                }

        if first == "week" and index + 1 < len(words):
            second = clean_text(words[index + 1].get("text", "")).lower()
            if second == "ending" and abs(words[index]["top"] - words[index + 1]["top"]) <= 3:
                week_header = {
                    "x0": min(words[index]["x0"], words[index + 1]["x0"]),
                    "x1": max(words[index]["x1"], words[index + 1]["x1"]),
                    "top": min(words[index]["top"], words[index + 1]["top"]),
                    "bottom": max(words[index]["bottom"], words[index + 1]["bottom"]),
                }

    return employee_header, week_header


def extract_employee_and_week_from_pdf(pdf_bytes):
    import pdfplumber

    employee_name = ""
    week_ending = None

    with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            words = page.extract_words(use_text_flow=False, keep_blank_chars=False)
            if not words:
                continue

            employee_header, week_header = find_header_positions(words)
            if not employee_header and not week_header:
                continue

            rows = group_words_into_rows(words, y_tolerance=3)
            if not rows:
                continue

            if employee_header and week_header:
                header_bottom = max(employee_header["bottom"], week_header["bottom"])
            elif employee_header:
                header_bottom = employee_header["bottom"]
            else:
                header_bottom = week_header["bottom"]

            data_row = None
            for row in rows:
                if row["top"] <= header_bottom + 2:
                    continue

                row_words = sorted(row["words"], key=lambda item: item["x0"])
                row_text = " ".join(clean_text(word["text"]) for word in row_words).strip().lower()

                if not row_text:
                    continue
                if "employee name" in row_text and "week ending" in row_text:
                    continue
                if row_text in ("employee", "name", "week", "ending"):
                    continue

                has_date = any(parse_date(clean_text(word["text"])) for word in row_words)
                if has_date:
                    data_row = row
                    break

            if not data_row:
                continue

            row_words = sorted(data_row["words"], key=lambda item: item["x0"])
            date_word = None

            for word in row_words:
                value = clean_text(word["text"])
                parsed = parse_date(value)
                if parsed:
                    week_ending = parsed
                    date_word = word
                    break

            name_words = []
            if week_header:
                name_right_boundary = week_header["x0"] - 5
            elif date_word:
                name_right_boundary = date_word["x0"] - 5
            else:
                name_right_boundary = float("inf")

            if employee_header:
                name_left_boundary = employee_header["x0"] - 20
            else:
                name_left_boundary = 0

            for word in row_words:
                value = clean_text(word["text"])
                if not value or parse_date(value):
                    continue

                center_x = (word["x0"] + word["x1"]) / 2
                if name_left_boundary <= center_x <= name_right_boundary:
                    name_words.append(word)

            if name_words:
                employee_name = clean_text(" ".join(word["text"] for word in name_words))

            if employee_name or week_ending:
                break

    return employee_name, week_ending


def extract_fields(text, pdf_bytes, filename):
    employee_name, week_ending = extract_employee_and_week_from_pdf(pdf_bytes)

    return {
        "Employee Name": employee_name,
        "Week Ending": week_ending,
        "Invoice Amount": parse_currency(extract_by_label(text, [r"Amount\s*Due"])),
        "System URN": parse_system_urn_from_filename(filename),
        "Invoice Number": clean_text(
            extract_by_label(text, [r"Invoice\s*Number", r"Invoice\s*No\.?"])
        ),
        "Invoice Date": parse_date(extract_by_label(text, [r"Invoice\s*Date"])),
        "Due Date": parse_date(extract_by_label(text, [r"Due\s*Date"])),
        "Project ID": clean_text(extract_by_label(text, [r"Project\s*ID"])),
    }


def set_header_and_width(ws):
    for column_index, header in enumerate(OUTPUT_COLUMNS, start=1):
        ws.cell(row=1, column=column_index, value=header)

    widths = {
        "A": 28,
        "B": 14,
        "C": 16,
        "D": 24,
        "E": 18,
        "F": 14,
        "G": 14,
        "H": 18,
    }

    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def apply_row_format(ws, row_num):
    for column in ["A", "D", "E", "H"]:
        cell = ws[f"{column}{row_num}"]
        cell.value = "" if cell.value is None else str(cell.value)
        cell.number_format = "@"

    for column in ["B", "F", "G"]:
        cell = ws[f"{column}{row_num}"]
        if cell.value:
            cell.number_format = "mm/dd/yyyy"

    invoice_amount = ws[f"C{row_num}"]
    if invoice_amount.value not in (None, ""):
        invoice_amount.number_format = "$#,##0.00"


def append_row(ws, row_data):
    ws.append(
        [
            row_data["Employee Name"],
            row_data["Week Ending"],
            row_data["Invoice Amount"],
            row_data["System URN"],
            row_data["Invoice Number"],
            row_data["Invoice Date"],
            row_data["Due Date"],
            row_data["Project ID"],
        ]
    )
    apply_row_format(ws, ws.max_row)


def worksheet_headers(ws):
    return [clean_text(ws.cell(row=1, column=index).value) for index in range(1, len(OUTPUT_COLUMNS) + 1)]


def prepare_worksheet(ws):
    headers = worksheet_headers(ws)
    if not any(headers):
        set_header_and_width(ws)
        return

    if headers != OUTPUT_COLUMNS:
        raise ValueError(
            "The uploaded Excel file does not match the expected MARC invoice template."
        )

    set_header_and_width(ws)


def create_workbook_bytes(rows, existing_workbook_bytes=None):
    from openpyxl import Workbook, load_workbook

    if existing_workbook_bytes:
        workbook = load_workbook(BytesIO(existing_workbook_bytes))
        worksheet = workbook.active
        prepare_worksheet(worksheet)
    else:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Invoices"
        set_header_and_width(worksheet)

    for row_data in rows:
        append_row(worksheet, row_data)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def process_pdf_file(uploaded_file):
    pdf_bytes = uploaded_file.getvalue()
    text = extract_text_from_pdf(pdf_bytes)
    if not text.strip():
        raise ValueError("No extractable text found in PDF.")
    return extract_fields(text, pdf_bytes, uploaded_file.name)


def process_uploaded_files(uploaded_files):
    rows = []
    errors = []

    for uploaded_file in uploaded_files:
        try:
            rows.append(process_pdf_file(uploaded_file))
        except Exception as exc:
            errors.append(f"{uploaded_file.name} -> {exc}")

    return rows, errors


def sanitize_output_filename(filename):
    candidate = clean_text(filename) or "marc_invoices.xlsx"
    if not candidate.lower().endswith(".xlsx"):
        candidate = f"{candidate}.xlsx"
    return re.sub(r'[<>:"/\\|?*]+', "_", candidate)


def build_input_signature(uploaded_pdfs, mode, existing_workbook):
    pdf_signature = tuple(
        (uploaded_file.name, len(uploaded_file.getvalue()))
        for uploaded_file in uploaded_pdfs
    )
    workbook_signature = None
    if existing_workbook is not None:
        workbook_signature = (existing_workbook.name, len(existing_workbook.getvalue()))
    return (pdf_signature, mode, workbook_signature)


def trigger_download(file_bytes, filename):
    payload = base64.b64encode(file_bytes).decode("utf-8")
    safe_name = filename.replace("\\", "_").replace('"', "_")
    components.html(
        f"""
        <html>
            <body>
                <script>
                    const link = document.createElement("a");
                    link.href = "data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{payload}";
                    link.download = "{safe_name}";
                    document.body.appendChild(link);
                    link.click();
                    link.remove();
                </script>
            </body>
        </html>
        """,
        height=0,
    )


def render_branding():
    logo_path = Path(__file__).with_name("Midea.png")
    if logo_path.exists():
        logo_base64 = base64.b64encode(logo_path.read_bytes()).decode("utf-8")
        st.markdown(
            f"""
            <div class="brand-shell">
                <div class="brand-logo-row">
                    <img class="brand-logo" src="data:image/png;base64,{logo_base64}" alt="Midea logo" />
                </div>
                <div class="brand-title-row">
                    <div class="brand-title">MARC Invoice Converter</div>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
    else:
        st.markdown(
            """
            <div class="brand-shell">
                <div class="brand-title-row">
                    <div class="brand-title">MARC Invoice Converter</div>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )


def render_app():
    st.set_page_config(
        page_title="MARC Invoice Converter",
        page_icon=":page_facing_up:",
        layout="centered",
    )

    st.markdown(
        """
        <style>
        .stApp {
            background:
                radial-gradient(circle at top right, rgba(0, 160, 233, 0.18), transparent 32%),
                radial-gradient(circle at top left, rgba(9, 70, 145, 0.10), transparent 28%),
                linear-gradient(180deg, #f5fbff 0%, #edf6fd 100%);
        }
        header[data-testid="stHeader"] {
            display: none;
        }
        .block-container {
            max-width: 920px;
            padding-top: 0.6rem;
            padding-bottom: 2rem;
        }
        .brand-shell {
            margin: 0 auto 1.15rem auto;
        }
        .brand-logo-row {
            display: flex;
            justify-content: flex-start;
            align-items: center;
            margin-bottom: 0.35rem;
        }
        .brand-title-row {
            display: flex;
            justify-content: center;
            text-align: center;
        }
        .brand-logo {
            height: 62px;
            width: auto;
            object-fit: contain;
        }
        .brand-title {
            color: #083b78;
            font-size: 2.2rem;
            font-weight: 800;
            line-height: 1.05;
            letter-spacing: -0.03em;
        }
        div[data-testid="stVerticalBlockBorderWrapper"] {
            border-radius: 24px;
            border: 1px solid rgba(0, 102, 179, 0.14);
            background: rgba(255, 255, 255, 0.90);
            box-shadow: 0 18px 45px rgba(17, 87, 151, 0.08);
        }
        div[data-testid="stFileUploader"] section {
            border: 1px dashed rgba(0, 118, 212, 0.32);
            border-radius: 18px;
            background: linear-gradient(180deg, rgba(233, 245, 255, 0.95) 0%, rgba(248, 252, 255, 0.96) 100%);
        }
        div[data-baseweb="radio"] > div {
            gap: 1rem;
        }
        .stButton > button, .stDownloadButton > button {
            border-radius: 999px;
            min-height: 3rem;
            font-weight: 700;
            border: none;
            background: linear-gradient(90deg, #0078d4 0%, #00a6e6 100%);
            color: white;
            box-shadow: 0 10px 24px rgba(0, 120, 212, 0.22);
        }
        .stButton > button:hover, .stDownloadButton > button:hover {
            background: linear-gradient(90deg, #006fc4 0%, #0098d3 100%);
            color: white;
        }
        .result-strip {
            margin-top: 1rem;
            padding: 1rem 1.15rem;
            border-radius: 18px;
            background: linear-gradient(90deg, rgba(0, 120, 212, 0.10) 0%, rgba(0, 166, 230, 0.06) 100%);
            border: 1px solid rgba(0, 120, 212, 0.12);
            color: #0b457d;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    render_branding()

    with st.container(border=True):
        uploaded_pdfs = st.file_uploader(
            "Invoice PDF files",
            type=["pdf"],
            accept_multiple_files=True,
        )

        mode = st.radio(
            "Workbook mode",
            options=["Create new workbook", "Append to existing workbook"],
            horizontal=True,
        )

        existing_workbook = None
        if mode == "Append to existing workbook":
            existing_workbook = st.file_uploader(
                "Existing Excel workbook",
                type=["xlsx"],
                accept_multiple_files=False,
            )

        output_name = st.text_input(
            "Downloaded file name",
            value="marc_invoices.xlsx",
        )

    if not uploaded_pdfs:
        return

    if mode == "Append to existing workbook" and existing_workbook is None:
        return

    input_signature = build_input_signature(uploaded_pdfs, mode, existing_workbook)
    generate_clicked = st.button("Generate workbook", type="primary", use_container_width=True)

    if generate_clicked:
        with st.spinner("Reading PDFs and extracting invoice data..."):
            rows, errors = process_uploaded_files(uploaded_pdfs)

        workbook_bytes = None
        workbook_error = None
        if rows:
            try:
                workbook_bytes = create_workbook_bytes(
                    rows,
                    existing_workbook.getvalue() if existing_workbook else None,
                )
            except Exception as exc:
                workbook_error = str(exc)

        st.session_state["invoice_results"] = {
            "signature": input_signature,
            "rows": rows,
            "errors": errors,
            "workbook_bytes": workbook_bytes,
            "workbook_error": workbook_error,
            "output_name": sanitize_output_filename(output_name),
            "auto_download": bool(workbook_bytes and not workbook_error),
        }

    results = st.session_state.get("invoice_results")
    if not results or results.get("signature") != input_signature:
        return

    rows = results["rows"]
    errors = results["errors"]
    workbook_bytes = results["workbook_bytes"]
    workbook_error = results.get("workbook_error")
    generated_output_name = results.get("output_name", sanitize_output_filename(output_name))

    if rows:
        st.markdown(
            (
                "<div class='result-strip'>"
                f"Processed {len(uploaded_pdfs)} PDF file(s), extracted {len(rows)} row(s)"
                f"{'' if not errors else f', with {len(errors)} file issue(s)'}."
                "</div>"
            ),
            unsafe_allow_html=True,
        )

        if workbook_error:
            st.error(f"Excel generation failed: {workbook_error}")
        elif workbook_bytes:
            if results.get("auto_download"):
                trigger_download(workbook_bytes, generated_output_name)
                st.session_state["invoice_results"]["auto_download"] = False
            st.success(f"Excel file is ready and download has started: {generated_output_name}")
    else:
        st.error("No valid invoice data was extracted from the uploaded PDFs.")

    if errors:
        with st.container(border=True):
            st.subheader("File issues")
            for error in errors:
                st.write(f"- {error}")


def main():
    render_app()


if __name__ == "__main__":
    main()
